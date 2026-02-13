import json
import os
import uuid
from copy import deepcopy
from datetime import datetime, timedelta

from flask import Flask, request, redirect, send_file, render_template_string, Response

from supabase import create_client
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_ANON_KEY")
CRM_TABLE = os.environ.get("CRM_TABLE", "crm_records")

SUPABASE = None
if SUPABASE_URL and SUPABASE_KEY:
    SUPABASE = create_client(SUPABASE_URL, SUPABASE_KEY)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

APP = Flask(__name__)

# ---------------- Supabase config ----------------
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_ANON_KEY")  # anon public
CRM_TABLE = os.environ.get("CRM_TABLE", "crm_records")
UNDO_TABLE = os.environ.get("UNDO_TABLE", "crm_undo")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise Exception("Faltan variables SUPABASE_URL o SUPABASE_ANON_KEY en Render")

SUPABASE = create_client(SUPABASE_URL, SUPABASE_KEY)
# ------------------------------------------------

EXPORT_FILE = "crm_export.xlsx"
UNDO_MAX = 30  # historial tipo Ctrl+Z

SERVICES = [
    "CEJAS",
    "DELINEADO DE OJOS",
    "LABIOS",
    "ACIDO HIALURONICO",
    "BIOESTIMULADORES",
    "RETOQUE",
]

COLUMNS_XLSX = ["NOMBRE", "TELEFONO", "FECHA", "FECHA RETOQUE", "SERVICIO", "COMENTARIO"]


# ---------- fechas ----------
def parse_ddmmyyyy(s: str):
    return datetime.strptime(s.strip(), "%d/%m/%Y")


def fmt_ddmmyyyy(dt: datetime):
    return dt.strftime("%d/%m/%Y")


def is_retouch_service(service: str) -> bool:
    return service.strip().upper() == "RETOQUE"


def compute_retouch_date(fecha_str: str, service: str) -> str:
    dt = parse_ddmmyyyy(fecha_str)
    days = 365 if is_retouch_service(service) else 20
    return fmt_ddmmyyyy(dt + timedelta(days=days))


def is_due(retouch_str: str) -> bool:
    try:
        r = parse_ddmmyyyy(retouch_str).date()
        return r <= datetime.now().date()
    except Exception:
        return False


# ---------- Supabase: datos ----------
def load_data():
    """Trae todos los registros desde Supabase."""
    try:
        res = (
            SUPABASE.table(CRM_TABLE)
            .select("id,nombre,telefono,fecha,servicio,comentario,recordatorio")
            .order("created_at", desc=True)
            .execute()
        )
        data = res.data or []
        out = []
        for r in data:
            out.append(
                {
                    "id": str(r.get("id") or uuid.uuid4()),
                    "nombre": (r.get("nombre") or "").strip(),
                    "telefono": (r.get("telefono") or "").strip(),
                    "fecha": (r.get("fecha") or "").strip(),  # dd/mm/yyyy
                    "servicio": (r.get("servicio") or "").strip(),
                    "comentario": (r.get("comentario") or "").strip(),
                    "recordatorio": bool(r.get("recordatorio", False)),
                }
            )
        return out
    except Exception:
        return []


def upsert_row(row: dict):
    """Crea o actualiza un registro en Supabase."""
    payload = {
        "id": row["id"],  # uuid string
        "nombre": row["nombre"],
        "telefono": row.get("telefono", "") or "",
        "fecha": row["fecha"],
        "servicio": row["servicio"],
        "comentario": row.get("comentario", "") or "",
        "recordatorio": bool(row.get("recordatorio", False)),
    }
    SUPABASE.table(CRM_TABLE).upsert(payload).execute()


def delete_row(rid: str):
    SUPABASE.table(CRM_TABLE).delete().eq("id", rid).execute()


# ---------- Supabase: Undo tipo Ctrl+Z ----------
def push_undo_snapshot(before_rows):
    """
    Guarda snapshot completo en Supabase.
    Mantiene máximo UNDO_MAX snapshots (borra los más antiguos).
    """
    try:
        SUPABASE.table(UNDO_TABLE).insert({"snapshot": before_rows}).execute()

        # mantener solo los últimos UNDO_MAX
        res = SUPABASE.table(UNDO_TABLE).select("id").order("id", desc=True).execute()
        ids = [r["id"] for r in (res.data or [])]
        if len(ids) > UNDO_MAX:
            to_delete = ids[UNDO_MAX:]
            SUPABASE.table(UNDO_TABLE).delete().in_("id", to_delete).execute()
    except Exception:
        pass


def pop_undo_snapshot():
    """Saca el último snapshot y lo elimina (LIFO)."""
    try:
        res = SUPABASE.table(UNDO_TABLE).select("id,snapshot").order("id", desc=True).limit(1).execute()
        rows = res.data or []
        if not rows:
            return None
        last = rows[0]
        SUPABASE.table(UNDO_TABLE).delete().eq("id", last["id"]).execute()
        return last.get("snapshot")
    except Exception:
        return None


def can_undo():
    try:
        res = SUPABASE.table(UNDO_TABLE).select("id").limit(1).execute()
        return bool(res.data)
    except Exception:
        return False


# ---------- validación ----------
def validate_row(nombre, fecha, servicio):
    if not nombre:
        return "El nombre es obligatorio."
    try:
        parse_ddmmyyyy(fecha)
    except Exception:
        return "Fecha inválida (selecciona una fecha del calendario)."
    if not servicio:
        return "El servicio es obligatorio."
    return None


HTML = r"""
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
  <meta name="theme-color" content="#2563eb"/>
  <link rel="manifest" href="/manifest.webmanifest">
  <title>CRM Salón</title>

  <style>
    body { font-family: Arial, sans-serif; margin: 14px; background:#f6f7fb; }
    .wrap { max-width: 1100px; margin: 0 auto; }
    .card { background: white; border-radius: 14px; padding: 14px; box-shadow: 0 6px 18px rgba(0,0,0,.06); margin-bottom: 12px; }
    h1 { margin: 0 0 10px 0; font-size: 20px; }
    .grid { display:grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .grid3 { display:grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }
    label { font-size: 12px; color: #333; }
    input, select, textarea { width: 100%; padding: 10px; border-radius: 10px; border: 1px solid #d9dbe7; font-size: 14px; background:#fff; }
    textarea { min-height: 70px; resize: vertical; }
    .row { display:flex; gap:10px; flex-wrap: wrap; align-items: center; }
    .btn { padding: 10px 12px; border-radius: 10px; border: 0; cursor:pointer; font-weight:700; text-decoration:none; display:inline-block; }
    .btn-primary { background:#2563eb; color:#fff; }
    .btn-danger { background:#ef4444; color:#fff; }
    .btn-ghost { background:#eef2ff; }
    .btn-ok { background:#16a34a; color:#fff; }
    .btn-warn { background:#f59e0b; color:#111827; }
    .btn-dark { background:#111827; color:#fff; }
    .muted { color:#555; font-size: 13px; }
    .banner { font-weight: 800; font-size: 14px; }
    .error { color:#b91c1c; font-weight:700; }

    table { width:100%; border-collapse: collapse; overflow:hidden; border-radius: 12px; }
    th, td { border: 1px solid #e6e8f2; padding: 10px; text-align: center; vertical-align: middle; font-size: 14px; }
    th { background:#d9ead3; font-size: 13px; }
    td.comment { text-align:left; white-space: pre-wrap; }
    tr.due { background: #fff2cc; }
    tr:hover { outline: 2px solid rgba(37,99,235,.15); }

    th.rem, td.rem { width: 62px; padding-left: 6px; padding-right: 6px; }
    th.act, td.act { width: 62px; padding-left: 6px; padding-right: 6px; }

    .chk { width: 22px; height: 22px; accent-color: #16a34a; cursor: pointer; }
    .chkWrap { display:flex; justify-content:center; align-items:center; }

    .tableWrap {
      overflow: auto;
      max-height: 60vh;
      border-radius: 12px;
      -webkit-overflow-scrolling: touch;
    }

    @media (max-width: 820px){
      .grid, .grid3 { grid-template-columns: 1fr; }
      th, td { font-size: 13px; padding: 8px; }
      th.rem, td.rem, th.act, td.act { width: 56px; }
      .tableWrap { max-height: 55vh; }
    }
  </style>
</head>

<body>
<div class="wrap">

  <div class="card">
    <h1>CRM Salón</h1>

    <div class="row" style="justify-content:space-between;">
      <div class="banner">
        {{ banner }}
        <br><span class="muted">Tip: toca una fila para cargarla arriba y editarla.</span>
      </div>
      <button class="btn btn-dark" type="button" onclick="goFullscreen()">⛶ Pantalla completa</button>
    </div>

    {% if error %}
      <p class="error">⚠️ {{ error }}</p>
    {% endif %}
  </div>

  <div class="card">
    <form method="post" action="/save" id="mainForm" autocomplete="off">
      <input type="hidden" name="id" id="rid" value="">

      <div class="grid">
        <div>
          <label>NOMBRE</label>
          <input name="nombre" id="nombre" placeholder="Ej: Romina" required autocomplete="off">
        </div>
        <div>
          <label>TELEFONO (opcional)</label>
          <input name="telefono" id="telefono" placeholder="Ej: 999999999 (opcional)" inputmode="tel" autocomplete="off">
        </div>
      </div>

      <div class="grid3" style="margin-top:10px;">
        <div>
          <label>FECHA (calendario)</label>
          <input type="hidden" name="fecha" id="fecha_hidden" value="{{ today_ddmmyyyy }}">
          <input id="fecha_picker" type="date" value="{{ today_iso }}" autocomplete="off">
        </div>

        <div>
          <label>SERVICIO</label>
          <select name="servicio" id="servicio" required autocomplete="off">
            <option value="">-- Selecciona --</option>
            {% for s in services %}
              <option value="{{ s }}">{{ s }}</option>
            {% endfor %}
          </select>
        </div>

        <div>
          <label>FECHA RETOQUE (auto)</label>
          <input id="retoque" readonly>
        </div>
      </div>

      <div style="margin-top:10px;">
        <label>COMENTARIO</label>
        <textarea name="comentario" id="comentario" placeholder="Ej: Se usó black y brown..." autocomplete="off"></textarea>
      </div>

      <div class="row" style="margin-top:12px;">
        <button class="btn btn-primary" type="submit">💾 Guardar (Agregar/Actualizar)</button>
        <button class="btn btn-ghost" type="button" onclick="clearForm()">🧹 Limpiar</button>

        <button class="btn btn-warn" type="submit"
                formaction="/undo" formmethod="post" formnovalidate
                {% if not can_undo %}disabled{% endif %}>
          ↩️ Deshacer
        </button>

        <a class="btn btn-ok" href="/export">📤 Exportar</a>
      </div>
    </form>
  </div>

  <div class="card">
    <div class="row" style="justify-content:space-between;">
      <div style="flex:1; min-width:240px;">
        <label>Filtrar (en vivo)</label>
        <input id="q" placeholder="Escribe para filtrar..." autocomplete="off">
        <div class="muted" style="margin-top:6px;">
          Filtra por: nombre, teléfono, servicio o comentario.
        </div>
      </div>
    </div>

    <div class="tableWrap" style="margin-top:10px;">
      <table id="crmTable">
        <thead>
          <tr>
            <th>NOMBRE</th>
            <th>TELEFONO</th>
            <th>FECHA</th>
            <th>RETOQUE</th>
            <th>SERVICIO</th>
            <th>COMENTARIO</th>
            <th class="rem">REC.</th>
            <th class="act">ACC.</th>
          </tr>
        </thead>

        <tbody id="crmBody">
          {% for r in rows %}
            <tr class="{{ r.row_class }}"
                data-search="{{ (r.nombre ~ ' ' ~ r.telefono ~ ' ' ~ r.fecha ~ ' ' ~ r.retoque ~ ' ' ~ r.servicio ~ ' ' ~ r.comentario)|lower }}"
                onclick='loadRow({{ r|tojson }})'
                style="cursor:pointer;">
              <td><b>{{ r.nombre }}</b></td>
              <td>{{ r.telefono }}</td>
              <td>{{ r.fecha }}</td>
              <td>{{ r.retoque }}</td>
              <td>{{ r.servicio }}</td>
              <td class="comment">{{ r.comentario }}</td>

              <td class="rem" onclick="event.stopPropagation();">
                <form method="post" action="/toggle_reminder" style="margin:0;" onclick="event.stopPropagation();">
                  <input type="hidden" name="id" value="{{ r.id }}">
                  <input type="hidden" name="target" value="{% if r.recordatorio %}0{% else %}1{% endif %}">
                  <div class="chkWrap">
                    <input class="chk" type="checkbox"
                           {% if r.recordatorio %}checked{% endif %}
                           onclick="event.stopPropagation();"
                           onchange="confirmReminder(this);">
                  </div>
                </form>
              </td>

              <td class="act" onclick="event.stopPropagation();">
                <form method="post" action="/delete"
                      onsubmit="return confirm('¿Eliminar este registro?');"
                      style="margin:0;" onclick="event.stopPropagation();">
                  <input type="hidden" name="id" value="{{ r.id }}">
                  <button class="btn btn-danger" type="submit" onclick="event.stopPropagation();">🗑️</button>
                </form>
              </td>
            </tr>
          {% endfor %}

          {% if rows|length == 0 %}
            <tr><td colspan="8" class="muted">No hay resultados.</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

</div>

<script>
  function goFullscreen(){
    const el = document.documentElement;
    if (el.requestFullscreen) el.requestFullscreen();
    else if (el.webkitRequestFullscreen) el.webkitRequestFullscreen();
  }

  function dateValueToDDMMYYYY(val){
    const parts = (val || "").split("-");
    if(parts.length !== 3) return "";
    const yy = parts[0], mm = parts[1], dd = parts[2];
    if(!yy || !mm || !dd) return "";
    return ${dd}/${mm}/${yy};
  }

  function ddmmyyyyToDateValue(ddmmyyyy){
    const parts = (ddmmyyyy || "").split("/");
    if(parts.length !== 3) return "";
    const dd = parts[0], mm = parts[1], yy = parts[2];
    if(dd.length!==2 || mm.length!==2 || yy.length!==4) return "";
    return ${yy}-${mm}-${dd};
  }

function computeRetouchFromHidden(servicio){
  const isRet = (servicio || "").trim().toUpperCase() === "RETOQUE";
  const days = isRet ? 365 : 21;

  // 1) intentamos desde el date picker (YYYY-MM-DD)
  const pickerVal = (document.getElementById("fecha_picker")?.value || "").trim();
  if (pickerVal) {
    const parts = pickerVal.split("-");
    if (parts.length === 3) {
      const y = parseInt(parts[0], 10);
      const m = parseInt(parts[1], 10) - 1;
      const d = parseInt(parts[2], 10);
      const dt = new Date(y, m, d);
      if (!isNaN(dt.getTime())) {
        dt.setDate(dt.getDate() + days);
        const dd = String(dt.getDate()).padStart(2,"0");
        const mm = String(dt.getMonth()+1).padStart(2,"0");
        const yy = dt.getFullYear();
        return ${dd}/${mm}/${yy};
      }
    }
  }

  // 2) fallback: hidden (DD/MM/YYYY o YYYY-MM-DD)
  const raw = (document.getElementById("fecha_hidden")?.value || "").trim();

  // si viene DD/MM/YYYY
  if (raw.includes("/")) {
    const p = raw.split("/");
    if (p.length === 3) {
      const d = parseInt(p[0],10), m = parseInt(p[1],10)-1, y = parseInt(p[2],10);
      const dt = new Date(y, m, d);
      if (!isNaN(dt.getTime())) {
        dt.setDate(dt.getDate() + days);
        const dd = String(dt.getDate()).padStart(2,"0");
        const mm = String(dt.getMonth()+1).padStart(2,"0");
        const yy = dt.getFullYear();
        return ${dd}/${mm}/${yy};
      }
    }
  }

  // si viene YYYY-MM-DD
  if (raw.includes("-")) {
    const p = raw.split("-");
    if (p.length === 3) {
      const y = parseInt(p[0],10), m = parseInt(p[1],10)-1, d = parseInt(p[2],10);
      const dt = new Date(y, m, d);
      if (!isNaN(dt.getTime())) {
        dt.setDate(dt.getDate() + days);
        const dd = String(dt.getDate()).padStart(2,"0");
        const mm = String(dt.getMonth()+1).padStart(2,"0");
        const yy = dt.getFullYear();
        return ${dd}/${mm}/${yy};
      }
    }
  }

  return "";
}
  }

  function updateRetouch(){
    const s = document.getElementById("servicio").value.trim();
    document.getElementById("retoque").value = computeRetouchFromHidden(s);
  }

  function syncHiddenFromPicker(){
    const picker = document.getElementById("fecha_picker");
    const hidden = document.getElementById("fecha_hidden");
    const ddmmyyyy = dateValueToDDMMYYYY(picker.value);
    if(ddmmyyyy) hidden.value = ddmmyyyy;
    updateRetouch();
  }

  function loadRow(r){
    document.getElementById("rid").value = r.id || "";
    document.getElementById("nombre").value = r.nombre || "";
    document.getElementById("telefono").value = r.telefono || "";
    document.getElementById("servicio").value = r.servicio || "";
    document.getElementById("comentario").value = r.comentario || "";

    const picker = document.getElementById("fecha_picker");
    const hidden = document.getElementById("fecha_hidden");
    hidden.value = r.fecha || hidden.value;
    const iso = ddmmyyyyToDateValue(hidden.value);
    if(iso) picker.value = iso;

    updateRetouch();
    window.scrollTo({ top: 0, behavior: "smooth" });
  }

  function clearForm(){
    document.getElementById("rid").value = "";
    document.getElementById("nombre").value = "";
    document.getElementById("telefono").value = "";
    document.getElementById("servicio").value = "";
    document.getElementById("comentario").value = "";

    document.getElementById("fecha_picker").value = "{{ today_iso }}";
    document.getElementById("fecha_hidden").value = "{{ today_ddmmyyyy }}";

    updateRetouch();
  }

  function liveFilter(){
    const q = (document.getElementById("q").value || "").trim().toLowerCase();
    const rows = document.querySelectorAll("#crmBody tr[data-search]");
    rows.forEach(tr => {
      const hay = (tr.getAttribute("data-search") || "");
      const show = !q || hay.includes(q);
      tr.style.display = show ? "" : "none";
    });
  }

  function confirmReminder(chk){
    const form = chk.closest("form");
    const hiddenTarget = form.querySelector('input[name="target"]');
    const want = chk.checked ? "1" : "0";
    hiddenTarget.value = want;

    const msg = (want === "1")
      ? "¿Se le envió recordatorio?\n\nSí = guardar marca"
      : "¿Quitar marca de recordatorio?\n\nSí = quitar";

    const ok = confirm(msg);
    if(ok){
      form.submit();
    } else {
      chk.checked = !chk.checked;
      hiddenTarget.value = chk.checked ? "1" : "0";
    }
  }

  document.getElementById("q").addEventListener("input", liveFilter);
  document.getElementById("fecha_picker").addEventListener("change", syncHiddenFromPicker);
  document.getElementById("servicio").addEventListener("change", updateRetouch);

  syncHiddenFromPicker();
  liveFilter();

  if ("serviceWorker" in navigator) {
    navigator.serviceWorker.register("/sw.js?v=12").catch(()=>{});
  }
alert("JS NUEVO CARGADO ✅");
</script>
</body>
</html>
"""


@APP.get("/manifest.webmanifest")
def manifest():
    m = {
        "name": "CRM Salon",
        "short_name": "CRM",
        "start_url": "/",
        "display": "fullscreen",
        "background_color": "#f6f7fb",
        "theme_color": "#2563eb",
        "icons": [],
    }
    return Response(json.dumps(m), mimetype="application/manifest+json")


@APP.get("/sw.js")
def sw():
    js = r"""
self.addEventListener("install", (e) => {
  self.skipWaiting();
});
self.addEventListener("activate", (e) => {
  e.waitUntil((async () => {
    try{
      const keys = await caches.keys();
      await Promise.all(keys.map(k => caches.delete(k)));
    }catch(e){}
    await self.clients.claim();
  })());
});
self.addEventListener("fetch", (e) => {
  e.respondWith(fetch(e.request).catch(() => caches.match(e.request)));
});
"""
    return Response(js, mimetype="application/javascript")


@APP.get("/")
def index():
    data = load_data()

    rows = []
    nearest = None

    for r in data:
        retoque = ""
        due = False
        try:
            retoque = compute_retouch_date(r["fecha"], r["servicio"])
            due = is_due(retoque)
        except Exception:
            pass

        if retoque:
            try:
                d = parse_ddmmyyyy(retoque).date()
                if nearest is None or d < nearest[0]:
                    nearest = (d, retoque)
            except Exception:
                pass

        rows.append({**r, "retoque": retoque, "row_class": ("due" if due else "")})

    today_iso = datetime.now().strftime("%Y-%m-%d")
    today_ddmmyyyy = datetime.now().strftime("%d/%m/%Y")

    banner = "Retoque: —"
    if nearest:
        d, txt = nearest
        diff = (d - datetime.now().date()).days
        if diff > 0:
            banner = f"Próximo retoque más cercano: {txt} (faltan {diff} días)"
        else:
            banner = f"⚠️ Hay retoques que ya tocan (ej: {txt})"

    error = request.args.get("error") or ""
    return render_template_string(
        HTML,
        rows=rows,
        services=SERVICES,
        today_iso=today_iso,
        today_ddmmyyyy=today_ddmmyyyy,
        banner=banner,
        error=error,
        can_undo=can_undo(),
    )


@APP.post("/save")
def save():
    if SUPABASE is None:
        return "ERROR: Supabase no inicializó. Revisa SUPABASE_URL y SUPABASE_ANON_KEY en Render.", 500

    rid = (request.form.get("id") or "").strip()
    nombre = (request.form.get("nombre") or "").strip()
    telefono = (request.form.get("telefono") or "").strip()
    fecha = (request.form.get("fecha") or "").strip()       # dd/mm/yyyy (tu hidden)
    servicio = (request.form.get("servicio") or "").strip()
    comentario = (request.form.get("comentario") or "").strip()

    err = validate_row(nombre, fecha, servicio)
    if err:
        return redirect(f"/?error={err}")

    # id
    if not rid:
        rid = str(uuid.uuid4())

    # convertir fecha dd/mm/yyyy -> yyyy-mm-dd (Supabase date)
    fecha_iso = parse_ddmmyyyy(fecha).strftime("%Y-%m-%d")

    payload = {
        "id": rid,
        "nombre": nombre,
        "telefono": telefono,
        "fecha": fecha_iso,
        "servicio": servicio,
        "comentario": comentario,
        "recordatorio": False,
    }

    SUPABASE.table(CRM_TABLE).upsert(payload).execute()
    return redirect("/")

@APP.post("/delete")
def delete():
    data = load_data()
    before = deepcopy(data)

    rid = (request.form.get("id") or "").strip()
    if rid:
        delete_row(rid)

    push_undo_snapshot(before)
    return redirect("/")


@APP.post("/toggle_reminder")
def toggle_reminder():
    data = load_data()
    before = deepcopy(data)

    rid = (request.form.get("id") or "").strip()
    target = (request.form.get("target") or "").strip()  # "1" marcar, "0" desmarcar
    want = True if target == "1" else False

    # buscar registro actual y actualizar
    for r in data:
        if r.get("id") == rid:
            r["recordatorio"] = want
            upsert_row(r)
            break

    push_undo_snapshot(before)
    return redirect("/")


@APP.post("/undo")
def undo():
    snap = pop_undo_snapshot()
    if snap is None:
        return redirect("/")

    # Reemplazar tabla completa con snapshot:
    # 1) borrar todo
    try:
        SUPABASE.table(CRM_TABLE).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    except Exception:
        pass

    # 2) reinsertar snapshot
    try:
        for r in (snap or []):
            # asegurar id
            if not r.get("id"):
                r["id"] = str(uuid.uuid4())
            upsert_row(r)
    except Exception:
        pass

    return redirect("/")


# ---------- Export Excel ----------
def build_export_excel(path: str, data: list):
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            if "CRM" in wb.sheetnames:
                ws = wb["CRM"]
                wb.remove(ws)
            ws = wb.create_sheet("CRM", 0)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "CRM"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "CRM"

    ws.append(COLUMNS_XLSX)

    for r in data:
        retoque = ""
        try:
            retoque = compute_retouch_date(r["fecha"], r["servicio"])
        except Exception:
            pass
        ws.append([r["nombre"], r["telefono"], r["fecha"], retoque, r["servicio"], r["comentario"]])

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, size=12)
    cell_font = Font(size=11)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    due_fill = PatternFill("solid", fgColor="FFF2CC")

    col_widths = {"A": 18, "B": 16, "C": 14, "D": 16, "E": 22, "F": 48}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 26
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 42

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(COLUMNS_XLSX)):
        row_idx = row[0].row
        retoque_cell = ws.cell(row=row_idx, column=4)
        due = (row_idx != 1) and is_due(str(retoque_cell.value or ""))

        for cell in row:
            cell.border = border
            cell.alignment = align
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = cell_font
                if due:
                    cell.fill = due_fill

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        sh = wb["Sheet"]
        if sh.max_row == 1 and sh.max_column == 1 and (sh["A1"].value is None):
            wb.remove(sh)

    wb.save(path)


@APP.get("/export")
def export():
    data = load_data()
    build_export_excel(EXPORT_FILE, data)
    return send_file(EXPORT_FILE, as_attachment=True, download_name=EXPORT_FILE)


if __name__ == "__main__":
    APP.run(host="0.0.0.0", port=5000, debug=True)